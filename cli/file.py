import argparse
import os
import sys
import json
import openpyxl
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from utils.excel import list_excel_files, load_sheet

def convert_excel(path: str):
    book = openpyxl.load_workbook(path)
    sheetnames = book.sheetnames
    sheets = []
    for i, name in enumerate(sheetnames):
        result = load_sheet(path, i)
        sheets.append({"name": name, "headers": result["headers"], "data": result["data"]})
    return {"file": os.path.basename(path), "sheets": sheets}

def write_json(path: str, payload: dict):
    base, _ = os.path.splitext(path)
    out = base + ".json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)

def run(dir_path: str):
    files = list_excel_files(dir_path)
    for item in files:
        p = item["path"]
        payload = convert_excel(p)
        write_json(p, payload)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dir", required=True)
    args = parser.parse_args()
    run(args.dir)

if __name__ == "__main__":
    main()
