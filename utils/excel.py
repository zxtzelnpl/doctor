import datetime
import os
import openpyxl
from io import BytesIO
from typing import TypedDict, NotRequired, List

def load_sheet(path: str, number: int = 0):
    key = (path, number)
    if not hasattr(load_sheet, '_cache'):
        load_sheet._cache = {}
    cache = load_sheet._cache
    if key in cache:
        return cache[key]

    book = openpyxl.load_workbook(path, number)
    sheetnames = book.sheetnames
    sheet = book[sheetnames[number]]

    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        data.append(row_dict)

    result = {
        "headers": headers,
        "data": data,
    }
    cache[key] = result
    return result

def export_sheet(title: str, headers: list, data: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    ws.append(headers)

    for record in data:
        row = [record.get(header, "") for header in headers]
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'导出{title}_{timestamp}.xlsx'

    return {
        "filename": filename,
        "output": output,
    }

class SheetSpec(TypedDict):
    path: str
    number: NotRequired[int]

def load_sheets(sheets: List[SheetSpec]):
    from concurrent.futures import ThreadPoolExecutor
    headers_all = []
    seen = set()
    data_all = []
    max_workers = min(len(sheets) or 1, max(1, (os.cpu_count() or 4)))
    work_items = [{"path": os.path.abspath(s["path"]), "number": s.get("number", 0)} for s in sheets]
    def _task(spec: SheetSpec):
        return load_sheet(spec["path"], spec.get("number", 0))
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = [ex.submit(_task, s) for s in work_items]
        results = []
        for fut in futures:
            try:
                results.append(fut.result())
            except Exception:
                results.append({"headers": [], "data": []})
    for result in results:
        headers = result["headers"]
        data = result["data"]
        for h in headers:
            if h not in seen:
                seen.add(h)
                headers_all.append(h)
        data_all.extend(data)
    return {
        "headers": headers_all,
        "data": data_all,
    }

def list_excel_files(base: str) -> List[SheetSpec]:
    res: List[SheetSpec] = []
    if os.path.isdir(base):
        # 递归遍历目录
        for root, _, files in os.walk(base):
            for name in files:
                if not name.lower().endswith('.xlsx'):
                    continue
                if name.startswith('~'):
                    continue
                # 保存相对路径，保持与_files_signature一致
                rel_path = os.path.relpath(os.path.join(root, name), base)
                res.append({"path": os.path.join(base, rel_path)})
    return res

def get_all_files_sheets(year: int | str, base: str = './files'):
    base_path = os.path.join(base, str(year))
    sheets = list_excel_files(base_path)
    return load_sheets(sheets)
